import os
from openpyxl import load_workbook, Workbook

script_dir = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join(script_dir, "devices.xlsx")


def load_devices_from_excel():
    devices = []
    print("File Path:", file_path)
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) == 4:
                hostname, ip_address, port, switch = row
                devices.append(
                    {
                        "hostname": hostname,
                        "ip_address": ip_address,
                        "port": port,
                        "switch": switch,
                    }
                )
            else:
                print(f"Invalid row: {row}")
    return devices


def save_device_to_excel(device):
    try:
        if not os.path.exists(file_path):
            wb = Workbook()
            ws = wb.active
            ws.append(["Hostname", "IP Address", "Port", "Switch"])
        else:
            wb = load_workbook(file_path)
            ws = wb.active

        ws.append(
            [device["hostname"], device["ip_address"], device["port"], device["switch"]]
        )
        wb.save(file_path)
        print("Device saved successfully.")
    except Exception as e:
        print("Error occurred while saving device:", e)
