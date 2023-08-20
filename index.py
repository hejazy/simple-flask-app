from flask import Flask, render_template, request, redirect, flash
from flask_wtf import FlaskForm
from wtforms import StringField, SubmitField
from wtforms.validators import DataRequired
import os
from openpyxl import load_workbook, Workbook

app = Flask(__name__)
app.secret_key = 'your_secret_key_here'

script_dir = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join(script_dir, 'devices.xlsx')

class DeviceForm(FlaskForm):
    hostname = StringField('Hostname', validators=[DataRequired()])
    ip_address = StringField('IP Address', validators=[DataRequired()])
    port = StringField('Port', validators=[DataRequired()])
    switch = StringField('Switch', validators=[DataRequired()])
    submit = SubmitField('Add Device')

# ... (Your previous code)

def load_devices_from_excel():
    devices = []
    print("File Path:", file_path)
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) == 4:
                hostname, ip_address, port, switch = row
                devices.append({'hostname': hostname, 'ip_address': ip_address, 'port': port, 'switch': switch})
            else:
                print(f"Invalid row: {row}")
    return devices

def save_device_to_excel(device):
    try:
        if not os.path.exists(file_path):
            wb = Workbook()
            ws = wb.active
            ws.append(['Hostname', 'IP Address', 'Port', 'Switch'])
        else:
            wb = load_workbook(file_path)
            ws = wb.active

        ws.append([device['hostname'], device['ip_address'], device['port'], device['switch']])
        wb.save(file_path)
        print("Device saved successfully.")
    except Exception as e:
        print("Error occurred while saving device:", e)


@app.route('/', methods=['GET', 'POST'])
def home():
    devices = load_devices_from_excel()
    search_query = request.args.get('search', '')
    if search_query:
        devices_filtered = [device for device in devices if search_query.lower() in device['hostname'].lower()]
        return render_template('home.html', devices=devices_filtered, search_query=search_query)
    return render_template('home.html', devices=devices, search_query='')

@app.route('/add_device', methods=['GET', 'POST'])
def add_device():
    form = DeviceForm()
    print(request.form)
    if request.method == 'POST' :
        hostname = request.form.get('hostname')
        ip_address = request.form.get('ip_address')
        port = request.form.get('port')
        switch = request.form.get('switch')
        device = {'hostname': hostname, 'ip_address': ip_address, 'port': port, 'switch': switch}
        save_device_to_excel(device)
        flash('Device added successfully.', 'success')
        return redirect('/')
    return render_template('add_device.html', form=form)

if __name__ == '__main__':
    app.run(debug=True)
